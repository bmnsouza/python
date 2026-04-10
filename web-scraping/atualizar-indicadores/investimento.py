import time
import logging

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException

# URL
URL_INVESTIDOR10_ETF = "https://investidor10.com.br/etfs-global"
URL_INVESTIDOR10_FII = "https://investidor10.com.br/fiis"

# HTML
DIV_1 = "/div[1]"
DIV_13 = "/div[13]"
DIV_BODY = "//div[@class='_card-body']"
DIV_COTACAO = "//div[@class='_card cotacao']"
DIV_DESC = "//div[@class='desc']"
DIV_DY = "//div[@class='_card dy']"
DIV_ETF = "/div[@class='etfCurrentQuotation']"
DIV_TABLE_INDICATORS = "//div[@id='table-indicators']"
DIV_VALUE = "//div[@class='value']"
SPAN_1 = "/span[1]"

# Outros
ARQUIVO_EXCEL = "Investimento.xlsx"
ABA_PLANILHA = "Indicadores"
ETF = "ETF"
FII = "FII"


def log_erro(mensagem):
    logger = logging.getLogger("meu_logger")

    if not logger.handlers:
        handler = logging.FileHandler("erro.log")
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        logger.setLevel(logging.ERROR)

    logger.error(mensagem)


def configurar_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    return webdriver.Chrome(options=options)


def converter_moeda(valor: str) -> float:
    return float(
        valor.replace("R$", "").replace("US$", "").replace(".", "").replace(",", ".").strip()
    )


def converter_percentual(valor: str) -> float:
    return float(valor.replace("%", "").replace(",", ".").strip()) / 100


def buscar_valor(driver, xpath: str) -> str | None:
    try:
        wait = WebDriverWait(driver, 1)
        return wait.until(EC.presence_of_element_located((By.XPATH, xpath))).text.strip()
    except (NoSuchElementException, TimeoutException):
        return None


def processar_ativos(driver, worksheet, url_base: str, inicio: int, tipo: str):
    print(f"\n*** {tipo.upper()} ***")
    row = inicio

    while True:
        ativo = worksheet.cell(row=row, column=2).value
        if not ativo:
            break

        driver.get(f"{url_base}/{ativo.lower()}/")

        try:
            if tipo == ETF:
                preco = buscar_valor(driver, DIV_COTACAO + DIV_BODY + DIV_ETF + SPAN_1)
                dividendo = buscar_valor(driver, DIV_DY + DIV_BODY + SPAN_1)

                if preco and dividendo:
                    print(f"{ativo} = Preço: {preco}; Dividendo: {dividendo}")
                    worksheet.cell(row=row, column=3).value = converter_moeda(preco)
                    worksheet.cell(row=row, column=4).value = converter_percentual(dividendo)
                else:
                    mensagem = f"Dados incompletos para {ativo}"
                    print(f"\n{mensagem}")
                    log_erro(mensagem)

            elif tipo == FII:
                preco = buscar_valor(driver, DIV_COTACAO + DIV_BODY + DIV_1 + SPAN_1)
                vp = buscar_valor(driver, DIV_TABLE_INDICATORS + DIV_13 + DIV_DESC + DIV_VALUE)
                dividendo = buscar_valor(driver, DIV_DY + DIV_BODY + DIV_1 + SPAN_1)

                if preco and vp and dividendo:
                    print(f"{ativo} = Preço: {preco}; VP: {vp}; Dividendo: {dividendo}")
                    worksheet.cell(row=row, column=3).value = converter_moeda(preco)
                    worksheet.cell(row=row, column=4).value = converter_moeda(vp)
                    worksheet.cell(row=row, column=5).value = converter_percentual(dividendo)
                else:
                    mensagem = f"Dados incompletos para {ativo}"
                    print(f"\n{mensagem}")
                    log_erro(mensagem)

        except Exception as e:
            mensagem = f"Erro ao buscar {tipo} {ativo}: {e}"
            print(f"\n{mensagem}")
            log_erro(mensagem)
            break

        row += 1


def main():
    inicio = time.time()
    driver = None
    workbook = None

    try:
        driver = configurar_driver()
        workbook = load_workbook(ARQUIVO_EXCEL)
        worksheet = workbook[ABA_PLANILHA]

        processar_ativos(driver, worksheet, URL_INVESTIDOR10_ETF, inicio=3, tipo=ETF)
        processar_ativos(driver, worksheet, URL_INVESTIDOR10_FII, inicio=10, tipo=FII)

        workbook.save(ARQUIVO_EXCEL)

        fim = time.time()
        duracao = fim - inicio
        print("\nPlanilha atualizada com sucesso")
        print(f"Tempo total: {duracao:.2f} segundos")
        input("\nPresssione ENTER para sair ...")
    except Exception as e:
        mensagem = f"Erro inesperado: {e}"
        print(f"\n{mensagem}")
        log_erro(mensagem)
    finally:
        if workbook:
            workbook.close()
        if driver:
            driver.quit()


if __name__ == "__main__":
    main()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time

# Configurações opcionais
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

# Inicializa o navegador
driver = webdriver.Chrome(options=options)
print("")

# Abre a planilha
excel_path = "atualizar-indicadores-fiis.xlsx"
workbook = load_workbook(excel_path)
worksheet = workbook["Indicadores"]

# Lê os ativos a partir da célula B3
row = 3
while True:
    ativo = worksheet.cell(row = row, column = 2).value
    if not ativo:
        break

    url = f"https://investidor10.com.br/fiis/{ativo.lower()}/"
    driver.get(url)
    time.sleep(1)

    try:
        # Obtêm os elementos Preço, VP e Dividendo
        preco = driver.find_element(By.XPATH, "//div[@class='_card cotacao']//div[@class='_card-body']/div[1]/span[1]").text.strip()
        vp = driver.find_element(By.XPATH, "//div[@id='table-indicators']/div[13]//div[@class='desc']//div[@class='value']").text.strip()
        dividendo = driver.find_element(By.XPATH, "//div[@class='_card dy']//div[@class='_card-body']/div[1]/span[1]").text.strip()
        
        # Imprime os valores dos elementos
        print(f"{ativo} = Preço: {preco}; VP: {vp}; Dividendo: {dividendo}")

        # Escreve na planilha
        worksheet.cell(row = row, column = 3).value = preco[3:len(preco)]
        worksheet.cell(row = row, column = 4).value = vp[3:len(vp)]
        worksheet.cell(row = row, column = 5).value = dividendo
    except Exception as e:        
        print(f"Erro ao buscar {ativo}: {e}")
        log = open("Log.txt", "w")
        log.write(str(e))
        log.close()
        break
    
    row += 1

# Salva as alterações
workbook.save(excel_path)
driver.quit()

print("")
print("Planilha atualizada com sucesso")
time.sleep(5)